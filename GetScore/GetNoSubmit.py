import openpyxl

excel_out_path = "C:\\Users\\jun\\Desktop\\作业情况\\.xlsx"

def getIndex(time, i):
    index = (time-1)*2 + i
    if time == 9:
        index += 1
    return index

if __name__ == '__main__':
    wb = openpyxl.load_workbook(excel_out_path)
    ws = wb.worksheets[0]
    myStuIds = []
    myStuNames = []
    noSubmitList = []
    flgs = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
    for i in range(2, 32):
        myStuIds.append(ws['A'+str(i)].value)
        myStuNames.append(ws['B'+str(i)].value)
        fvalue = []
        for a in flgs:
            fvalue.append(ws[a+str(i)].value)
        noSubmitList.append(fvalue)
    
    for time in range(4, 10):
        excel_path = "C:\\Users\\jun\\Desktop\\作业情况\\a"+str(time)+".xlsx"
        wb1 = openpyxl.load_workbook(excel_path)
        ws1 = wb1.worksheets[0]
        pros = ['D', 'F', 'H', 'J']
        for i in range(1, 115):
            if myStuIds.__contains__(ws1['A'+str(i)].value):
                for j in range(4):
                    if "未提交" == ws1[pros[j] + str(i)].value or "未交" == ws1[pros[j] + str(i)].value:
                        index = myStuIds.index(ws1['A'+str(i)].value)
                        temp = noSubmitList[index]
                        temp[getIndex(time, j)] = '+'
                        noSubmitList[index] = temp
    
    for i in range(2, 32):
        for j in range(len(flgs)):
            ws[flgs[j]+str(i)] = noSubmitList[i-2][j]
    wb.save(excel_out_path)