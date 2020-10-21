from bs4 import BeautifulSoup
import requests
import openpyxl

# 学生成绩所在的网址url
data_url = "http://114.212.86.242:10002/index.php/submissions/final/problem/1"
# Excel路径，将学生成绩写到此excel表中
excel_path = "C:\\Users\\jun\\Desktop\\2作业2020-p4.xlsx"
# 问题几，会影响写到excel表中的哪一列
probNum = 1

login_url = 'http://114.212.86.242:10002/index.php/login'
header = {
    'Host': '114.212.86.242:10002',
    'Origin': 'http://114.212.86.242:10002',
    'Referer': login_url,
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.69 Safari/537.36 Edg/81.0.416.34"
}


# 使用requests模拟登录
def login(username, password):
    # 获取合法token
    _session_requests = requests.session()
    _r = _session_requests.get(login_url)
    soup = BeautifulSoup(_r.text, 'html.parser')
    token = soup.find(name="input", attrs={'name': 'shj_csrf_token', 'type': 'hidden'})
    if token is None:
        raise Exception("connecting error!")
    post_data = {
        'shj_csrf_token': token.attrs.get('value'),
        'username': username,
        'password': password
    }
    _session_requests.post(login_url, data=post_data, headers=header)
    return _session_requests


if __name__ == '__main__':
    # 登录oj
    session_requests = login('mf1933101', '19961211')
    # 获取网页文档
    r = session_requests.get(data_url, headers=header)
    # 使用BeautifulSoup库解析网页
    mySoup = BeautifulSoup(r.text, 'html.parser')
    # 每一行保存一个学生的信息
    trs = mySoup.find_all(name='tr')
    stuInfo = {}
    for tr in trs:
        if tr.attrs.get('data-u') is not None:
            tds = tr.find_all('td')
            if tds[2].text[0]=='1' or tds[2].text[0]=='2' : # 过滤非本科生的提交记录
                delay = tds[7].text.strip()
                stuInfo[tds[2].text] = [int(tds[8].text.strip())/10, int(delay[delay.index('\n')+1:delay.index('%')].strip())]
    print('submitted_total_count:', len(stuInfo))

    # 将已提交学生的信息写入指定Excel中
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[0]
    W1 = chr(ord('C') + (probNum-1)*2)
    W2 = chr(ord(W1)+1)
    # 跳过第一行的标题，ws.max_row为最大行
    for i in range(2, ws.max_row):
        stu_id = ws['A' + str(i)].value # excel表第一列为学号
        if stu_id is not None:
            if stuInfo.get(str(stu_id)) is not None:
                ws[W1 + str(i)] = stuInfo.get(str(stu_id))[0] # 通过学号获取学生成绩
                ws[W2 + str(i)] = ''
                if stuInfo.get(str(stu_id))[1] != 100:
                    ws[W2 + str(i)] = '延迟提交'
            else:
                ws[W1 + str(i)] = 0
                ws[W2 + str(i)] = '未提交'
    wb.save(excel_path)